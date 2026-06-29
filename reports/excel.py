"""
Excel Reporter — Azure Audit Pro
Exact same format/structure as AWS CloudAudit Pro excel.py.
Colors: Critical=Dark Red, High=Light Red, Medium=Yellow, Low=Green
Compliance columns populated in EVERY finding row.
"""
from pathlib import Path
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from engine.compliance_calc import calculate_compliance, severity_counts


class ExcelReporter:

    # ── Severity colors — identical to AWS tool ─────────────────────────────
    SEV = {
        "Critical": {"bg": "C00000", "fg": "FFFFFF"},  # Dark Red
        "High":     {"bg": "FF0000", "fg": "FFFFFF"},  # Light Red
        "Medium":   {"bg": "FFFF00", "fg": "000000"},  # Yellow
        "Low":      {"bg": "92D050", "fg": "000000"},  # Green
    }

    C = {
        "hdr_dark":   "1F3864", "hdr_dark_fg":  "FFFFFF",
        "hdr_med":    "0078D4", "hdr_med_fg":   "FFFFFF",   # Azure blue
        "hdr_light":  "D6E4F0", "hdr_light_fg": "1F3864",
        "hdr_azure":  "0078D4", "hdr_azure_fg": "FFFFFF",
        "white":      "FFFFFF", "light_gray":   "F5F5F5",
        "border":     "BFBFBF",
        "path_bg":    "EBF3FB", "path_fg":      "1F3864",
        "issue_bg":   "FFF0F0", "issue_fg":     "5C2D00",
        "rem_bg":     "F0FFF4", "rem_fg":       "1D4D1D",
        "ct_bg":      "FFFDE7", "ct_fg":        "5D4037",
        "tag_bg":     "F3E5F5", "tag_fg":       "4A148C",
        "pass_bg":    "E2EFDA", "pass_fg":      "375623",
        "warn_bg":    "FFE699", "warn_fg":      "7D4F00",
        "fail_bg":    "FFCCCC", "fail_fg":      "C00000",
    }

    # ── Column definitions — same structure as AWS, Azure-adapted ───────────
    COLS = [
        ("Finding ID",         14),
        ("Resource ID",        36),
        ("Resource Name",      18),
        ("Resource Type",      16),
        ("Resource Group",     20),
        ("Location",           14),
        ("Check Name",         36),
        ("Severity",           12),
        # Compliance columns
        ("CIS",                14),
        ("NIST CSF",           12),
        ("NIST 800-53",        14),
        ("CSA CCM",            12),
        ("MITRE ATT&CK",       16),
        ("HIPAA",              16),
        ("PCI DSS",            12),
        ("GDPR",               10),
        # Detail columns
        ("Exact Path",         52),
        ("Issue Detail",       58),
        ("Tags",               28),
        ("Owner Tag",          20),
        ("Subscription",       32),
        ("Remediation",        72),
    ]

    def __init__(self, display):
        self.display = display

    def generate(self, findings, out_dir, account_name, company, partial=False):
        wb         = Workbook()
        by_service = {}
        for f in findings:
            by_service.setdefault(f.get("service", "Other"), []).append(f)

        compliance = calculate_compliance(findings)
        sev        = severity_counts(findings)

        self._sheet_summary(wb, findings, account_name, company, compliance, sev, partial)
        self._sheet_stats(wb, findings, by_service, compliance)

        critical = [f for f in findings if f.get("severity") == "Critical"]
        if critical:
            self._sheet_findings(wb, "Critical Findings", "C00000", critical)

        for svc, svc_f in by_service.items():
            fails = [f for f in svc_f if f.get("status") == "FAIL"]
            if fails:
                self._sheet_findings(wb, svc[:31], self.C["hdr_dark"], fails)

        suffix = "_PARTIAL" if partial else ""
        fpath  = Path(out_dir) / f"{account_name}{suffix}.xlsx"
        wb.save(str(fpath))
        self.display.success(f"  Excel → {fpath}")
        return fpath

    # ── Helpers ───────────────────────────────────────────────────────────────

    def _fill(self, hex_c):
        return PatternFill("solid", fgColor=hex_c)

    def _font(self, bold=False, color="000000", size=9, italic=False):
        return Font(bold=bold, color=color, size=size, italic=italic, name="Calibri")

    def _border(self):
        s = Side(style="thin", color=self.C["border"])
        return Border(left=s, right=s, top=s, bottom=s)

    def _cell(self, ws, ref, val, bg=None, fg="000000", bold=False,
               halign="left", sz=9, italic=False, wrap=True):
        c = ws[ref]
        c.value     = val
        c.font      = self._font(bold=bold, color=fg, size=sz, italic=italic)
        c.fill      = self._fill(bg) if bg else self._fill(self.C["white"])
        c.alignment = Alignment(horizontal=halign, vertical="center", wrap_text=wrap)
        c.border    = self._border()

    def _hdr(self, ws, ref, text, bg, fg="FFFFFF", sz=9):
        self._cell(ws, ref, text, bg, fg, bold=True, halign="center", sz=sz)

    def _compliance_color(self, pct):
        if pct >= 80:   return self.C["pass_bg"], self.C["pass_fg"]
        elif pct >= 60: return self.C["warn_bg"], self.C["warn_fg"]
        else:           return self.C["fail_bg"], self.C["fail_fg"]

    # ── Extract compliance values from finding ────────────────────────────────

    def _extract_compliance(self, f):
        """Extract compliance requirement IDs per framework — same as AWS base."""
        comp = f.get("compliance", [])
        result = {
            "CIS": [], "NIST_CSF": [], "NIST_800_53": [],
            "CSA_CCM": [], "MITRE_ATTACK": [], "HIPAA": [],
            "PCI_DSS": [], "GDPR": [],
        }
        for c in comp:
            fw  = c.get("framework", "")
            req = c.get("requirement", "")
            if fw in result and req:
                result[fw].append(req)
        return {
            "CIS":      ", ".join(result["CIS"])          or "—",
            "NIST_CSF": ", ".join(result["NIST_CSF"])     or "—",
            "N800":     ", ".join(result["NIST_800_53"])  or "—",
            "CSA":      ", ".join(result["CSA_CCM"])      or "—",
            "MITRE":    ", ".join(result["MITRE_ATTACK"]) or "—",
            "HIPAA":    ", ".join(result["HIPAA"])        or "—",
            "PCI":      ", ".join(result["PCI_DSS"])      or "—",
            "GDPR":     ", ".join(result["GDPR"])         or "—",
        }

    # ── Executive Summary ─────────────────────────────────────────────────────

    def _sheet_summary(self, wb, findings, account_name, company,
                       compliance, sev, partial):
        ws = wb.active
        ws.title = "Executive Summary"
        ws.sheet_view.showGridLines = False
        ws.column_dimensions["A"].width = 36
        ws.column_dimensions["B"].width = 42

        suffix = " [PARTIAL SCAN]" if partial else ""
        ws.merge_cells("A1:B1")
        self._cell(ws, "A1",
            f"🛡  AZURE CLOUD SECURITY ASSESSMENT REPORT{suffix}",
            self.C["hdr_dark"], self.C["hdr_dark_fg"], bold=True, halign="center", sz=14)
        ws.row_dimensions[1].height = 38

        ws.merge_cells("A2:B2")
        self._cell(ws, "A2",
            "Provider: Azure  ·  Frameworks: CIS · NIST CSF · NIST 800-53 · "
            "CSA CCM · MITRE · HIPAA · PCI DSS · GDPR",
            self.C["hdr_azure"], self.C["hdr_azure_fg"], bold=True, halign="center", sz=9)
        ws.row_dimensions[2].height = 20

        meta = [
            ("Company",               company),
            ("Subscription",          account_name),
            ("Scan Date",             datetime.now().strftime("%Y-%m-%d %H:%M:%S")),
            ("Total Vulnerabilities", sev["Total"]),
        ]
        for i, (k, v) in enumerate(meta, start=3):
            ws.row_dimensions[i].height = 20
            self._cell(ws, f"A{i}", k, self.C["hdr_light"], self.C["hdr_light_fg"], bold=True)
            self._cell(ws, f"B{i}", v)

        # Severity summary
        r = len(meta) + 4
        ws.merge_cells(f"A{r}:B{r}")
        self._cell(ws, f"A{r}", "VULNERABILITY SUMMARY",
                   self.C["hdr_med"], self.C["hdr_med_fg"], bold=True, halign="center", sz=11)
        ws.row_dimensions[r].height = 26

        sev_rows = [
            ("🔴 Critical", sev["Critical"], self.SEV["Critical"]["bg"], self.SEV["Critical"]["fg"]),
            ("🟠 High",     sev["High"],     self.SEV["High"]["bg"],     self.SEV["High"]["fg"]),
            ("🟡 Medium",   sev["Medium"],   self.SEV["Medium"]["bg"],   self.SEV["Medium"]["fg"]),
            ("🟢 Low",      sev["Low"],      self.SEV["Low"]["bg"],      self.SEV["Low"]["fg"]),
            ("Total",       sev["Total"],    self.C["hdr_light"],        self.C["hdr_light_fg"]),
        ]
        for i, (label, val, bg, fg) in enumerate(sev_rows, start=r + 1):
            ws.row_dimensions[i].height = 22
            self._cell(ws, f"A{i}", label, bg, fg, bold=True)
            self._cell(ws, f"B{i}", val,   bg, fg, bold=True, halign="center")

        # Framework compliance
        fw_r = r + len(sev_rows) + 2
        ws.merge_cells(f"A{fw_r}:B{fw_r}")
        self._cell(ws, f"A{fw_r}",
            "FRAMEWORK COMPLIANCE  (calculated from actual FAIL findings)",
            self.C["hdr_med"], self.C["hdr_med_fg"], bold=True, halign="center", sz=10)
        ws.row_dimensions[fw_r].height = 26

        for i, (fw, pct) in enumerate(sorted(compliance.items()), start=fw_r + 1):
            ws.row_dimensions[i].height = 20
            bg, fg = self._compliance_color(pct)
            self._cell(ws, f"A{i}", fw,          self.C["hdr_light"], self.C["hdr_light_fg"], bold=True)
            self._cell(ws, f"B{i}", f"{pct}%", bg, fg, bold=True, halign="center")

        # Developer note
        note_r = fw_r + len(compliance) + 2
        ws.merge_cells(f"A{note_r}:B{note_r}")
        self._cell(ws, f"A{note_r}",
            "ℹ  Azure Audit Pro v1  ·  Developed by 🦅 Singaram  ·  "
            "20 Azure Services · 120+ Checks · 8 Compliance Frameworks",
            self.C["ct_bg"], self.C["ct_fg"], italic=True, sz=8)
        ws.row_dimensions[note_r].height = 36

    # ── Risk & Statistics ─────────────────────────────────────────────────────

    def _sheet_stats(self, wb, findings, by_service, compliance):
        ws = wb.create_sheet("Risk & Statistics")
        ws.sheet_view.showGridLines = False

        ws.merge_cells("A1:G1")
        self._cell(ws, "A1", "RISK SCORE & STATISTICS",
                   self.C["hdr_dark"], self.C["hdr_dark_fg"], bold=True, halign="center", sz=12)
        ws.row_dimensions[1].height = 30

        # Framework compliance table
        ws.merge_cells("A3:G3")
        self._cell(ws, "A3",
            "Framework Compliance % — calculated from actual FAIL findings",
            self.C["hdr_med"], self.C["hdr_med_fg"], bold=True, halign="center", sz=10)

        fw_cols   = ["Framework", "Total Requirements", "Failed", "Passed", "Compliance %", "Status"]
        fw_widths = [30, 20, 14, 14, 14, 14]
        for ci, (col, w) in enumerate(zip(fw_cols, fw_widths), 1):
            ws.column_dimensions[get_column_letter(ci)].width = w
            self._hdr(ws, f"{get_column_letter(ci)}4", col, self.C["hdr_dark"])

        from engine.compliance_calc import load_compliance_requirements
        fw_reqs       = load_compliance_requirements("azure")
        failed_checks = {f.get("check_id") for f in findings if f.get("status") == "FAIL"}
        display_names = {
            "CIS":          "CIS Azure Benchmark 2.0",
            "NIST_CSF":     "NIST CSF 2.0",
            "NIST_800_53":  "NIST 800-53 Rev5",
            "CSA_CCM":      "CSA CCM v4.0",
            "MITRE_ATTACK": "MITRE ATT&CK Cloud",
            "HIPAA":        "HIPAA Security Rule",
            "PCI_DSS":      "PCI DSS 4.0",
            "GDPR":         "GDPR",
        }

        for ri, (fw_key, reqs) in enumerate(fw_reqs.items(), start=5):
            ws.row_dimensions[ri].height = 20
            total  = len(reqs)
            failed = sum(1 for req in reqs
                         if any(c in failed_checks for c in req.get("Checks", [])))
            passed = total - failed
            pct    = round((passed / total * 100), 1) if total else 0.0
            bg, fg = self._compliance_color(pct)
            status = "✅ Good" if pct >= 80 else "⚠ Needs Work" if pct >= 60 else "❌ Poor"
            row    = [display_names.get(fw_key, fw_key), total, failed, passed, f"{pct}%", status]
            for ci, val in enumerate(row, 1):
                c_bg = self.C["hdr_light"] if ci == 1 else (bg if ci >= 5 else
                        (self.C["light_gray"] if ri % 2 == 0 else self.C["white"]))
                c_fg = self.C["hdr_light_fg"] if ci == 1 else (fg if ci >= 5 else "000000")
                self._cell(ws, f"{get_column_letter(ci)}{ri}", val, c_bg, c_fg,
                           bold=(ci == 1 or ci >= 5),
                           halign="center" if ci > 1 else "left")

        # Per-service stats
        sec2 = len(fw_reqs) + 7
        ws.merge_cells(f"A{sec2}:G{sec2}")
        self._cell(ws, f"A{sec2}", "Findings Per Azure Service",
                   self.C["hdr_med"], self.C["hdr_med_fg"], bold=True, halign="center", sz=10)

        svc_cols = ["Service", "Critical", "High", "Medium", "Low", "Total", "Top Risk"]
        svc_ws   = [24, 12, 12, 12, 12, 12, 44]
        for ci, (col, w) in enumerate(zip(svc_cols, svc_ws), 1):
            ws.column_dimensions[get_column_letter(ci)].width = max(
                w, ws.column_dimensions[get_column_letter(ci)].width)
            self._hdr(ws, f"{get_column_letter(ci)}{sec2 + 1}", col, self.C["hdr_dark"])

        for ri, (svc, svc_f) in enumerate(by_service.items(), start=sec2 + 2):
            ws.row_dimensions[ri].height = 20
            fails = [f for f in svc_f if f.get("status") == "FAIL"]
            crit  = sum(1 for f in fails if f.get("severity") == "Critical")
            high  = sum(1 for f in fails if f.get("severity") == "High")
            med   = sum(1 for f in fails if f.get("severity") == "Medium")
            low   = sum(1 for f in fails if f.get("severity") == "Low")
            top   = fails[0].get("check_name", "")[:55] if fails else "—"
            bg    = self.C["light_gray"] if ri % 2 == 0 else self.C["white"]
            for ci, val in enumerate([svc, crit, high, med, low, len(fails), top], 1):
                self._cell(ws, f"{get_column_letter(ci)}{ri}", val,
                           self.C["hdr_light"] if ci == 1 else bg,
                           self.C["hdr_light_fg"] if ci == 1 else "000000",
                           bold=(ci == 1), halign="center" if 1 < ci < 7 else "left")

    # ── Findings sheet ────────────────────────────────────────────────────────

    def _sheet_findings(self, wb, title, color, findings):
        ws     = wb.create_sheet(title[:31])
        ws.sheet_view.showGridLines = False
        ncols  = len(self.COLS)
        last_c = get_column_letter(ncols)

        # Title banner
        ws.merge_cells(f"A1:{last_c}1")
        self._cell(ws, "A1",
            f"🛡  {title} — Azure Security Findings (FAIL Only)",
            color, "FFFFFF", bold=True, halign="center", sz=11)
        ws.row_dimensions[1].height = 30

        # Sub-header with counts
        ws.merge_cells(f"A2:{last_c}2")
        c = {s: sum(1 for f in findings if f.get("severity") == s)
             for s in ("Critical", "High", "Medium", "Low")}
        self._cell(ws, "A2",
            f"Total: {len(findings)}  |  🔴 Critical: {c['Critical']}  |  "
            f"🟠 High: {c['High']}  |  🟡 Medium: {c['Medium']}  |  🟢 Low: {c['Low']}  |  "
            f"🦅 Azure Audit Pro v1 — Developed by Singaram",
            self.C["hdr_med"], "FFFFFF", sz=8, italic=True, halign="left")
        ws.row_dimensions[2].height = 18

        # Column headers
        for ci, (cname, cw) in enumerate(self.COLS, 1):
            ws.column_dimensions[get_column_letter(ci)].width = cw
            self._hdr(ws, f"{get_column_letter(ci)}3", cname, self.C["hdr_dark"])
        ws.row_dimensions[3].height = 28
        ws.freeze_panes = "A4"
        ws.auto_filter.ref = f"A3:{last_c}3"

        # Finding rows
        for ri, f in enumerate(findings, start=4):
            ws.row_dimensions[ri].height = 80
            sev     = f.get("severity", "Low")
            sev_cfg = self.SEV.get(sev, {"bg": "FFFFFF", "fg": "000000"})
            row_bg  = sev_cfg["bg"]
            sev_fg  = sev_cfg["fg"]
            alt_bg  = self.C["light_gray"] if ri % 2 == 0 else self.C["white"]

            comp = self._extract_compliance(f)
            fid  = f"FIND-{f.get('service','')}-{str(ri-3).zfill(3)}"

            row_vals = [
                fid,
                f.get("resource_id",    ""),
                f.get("resource_name",  ""),
                f.get("resource_type",  ""),
                f.get("resource_group", ""),
                f.get("location",       ""),
                f.get("check_name",     ""),
                sev,
                # ── Compliance columns ──
                comp["CIS"],
                comp["NIST_CSF"],
                comp["N800"],
                comp["CSA"],
                comp["MITRE"],
                comp["HIPAA"],
                comp["PCI"],
                comp["GDPR"],
                # ── Detail columns ──
                f.get("exact_path",           ""),
                f.get("issue_detail",         ""),
                f.get("tags",                 ""),
                f.get("owner_tag",            ""),
                f.get("subscription_alias",   ""),
                f.get("remediation",          ""),
            ]

            for ci, val in enumerate(row_vals, 1):
                ref = f"{get_column_letter(ci)}{ri}"

                if ci == 8:           # Severity
                    self._cell(ws, ref, val, row_bg, sev_fg, bold=True, halign="center")
                elif 9 <= ci <= 16:   # Compliance cols
                    has_val = val and val != "—"
                    cell_bg = self.C["hdr_light"] if has_val else alt_bg
                    cell_fg = self.C["hdr_light_fg"] if has_val else "9E9E9E"
                    self._cell(ws, ref, val, cell_bg, cell_fg, sz=8, halign="center")
                elif ci == 17:        # Exact Path
                    self._cell(ws, ref, val, self.C["path_bg"], self.C["path_fg"], sz=8)
                elif ci == 18:        # Issue Detail
                    self._cell(ws, ref, val, self.C["issue_bg"], self.C["issue_fg"], sz=8)
                elif ci in (19, 20):  # Tags, Owner
                    self._cell(ws, ref, val, self.C["tag_bg"], self.C["tag_fg"], sz=8, italic=True)
                elif ci == 21:        # Subscription
                    self._cell(ws, ref, val, self.C["ct_bg"], self.C["ct_fg"], sz=9)
                elif ci == 22:        # Remediation
                    self._cell(ws, ref, val, self.C["rem_bg"], self.C["rem_fg"], sz=9)
                elif ci == 1:         # Finding ID
                    self._cell(ws, ref, val, self.C["hdr_light"], self.C["hdr_light_fg"],
                               bold=True, sz=8)
                else:
                    self._cell(ws, ref, val, alt_bg, "000000", sz=9)


# ── Standalone function wrapper (called from main.py) ─────────────────────────

def generate_excel_report(all_findings, subscriptions_info, compliance_results,
                           sev_counts, output_dir, display):
    import os
    os.makedirs(output_dir, exist_ok=True)
    reporter     = ExcelReporter(display)
    sub_alias    = (subscriptions_info[0].get("subscription_alias", "azure-audit")
                    if subscriptions_info else "azure-audit")
    company      = "Azure Security Assessment"
    path         = reporter.generate(
        [f for f in all_findings if f.get("status") == "FAIL"],
        output_dir, sub_alias, company
    )
    return str(path)
